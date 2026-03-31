import re
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

_WS_RE = re.compile(r"\s+")


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return _WS_RE.sub(" ", str(value).strip().casefold())


# Заголовки как в «Шаблон проектное задание 1.xlsx»; столбцы договор/этап в файле игнорируются — берутся из карточки проекта.
_HEADER_EXACT: Dict[str, str] = {
    "наименование задачи": "task_name",
    "описание задачи": "task_adenda",
    "исполнитель": "worker_name",
    "дата начала задачи": "task_start_date",
    "дата окончания задачи": "task_end_date",
    "комментарий": "task_comment",
    "рабочий файл": "working_file",
    "статус задачи": "task_status",
}

_SKIP_HEADERS = {"номер договора", "номер этапа"}


def _map_header_to_field(h_norm: str) -> Optional[str]:
    if not h_norm or h_norm in _SKIP_HEADERS:
        return None
    if h_norm in _HEADER_EXACT:
        return _HEADER_EXACT[h_norm]
    # Допускаем укороченные подписи
    if "наименование" in h_norm and "задач" in h_norm:
        return "task_name"
    if "описание" in h_norm and "задач" in h_norm:
        return "task_adenda"
    if h_norm == "исполнитель":
        return "worker_name"
    if "дата начала" in h_norm:
        return "task_start_date"
    if "дата окончания" in h_norm or "окончания задачи" in h_norm:
        return "task_end_date"
    if h_norm == "комментарий":
        return "task_comment"
    if "рабочий" in h_norm and "файл" in h_norm:
        return "working_file"
    if "статус" in h_norm:
        return "task_status"
    return None


def _build_column_map_from_header_row(row_vals: List[Any]) -> Dict[str, int]:
    """field_name -> 1-based column index."""
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(row_vals, start=1):
        field = _map_header_to_field(_norm_header(cell))
        if field and field not in mapping:
            mapping[field] = idx
    return mapping


def _parse_row_by_positions(
    row_vals: List[Any],
    *,
    start_col: int,
    contract_number: str,
    etap_number: str,
) -> Dict[str, Any]:
    """start_col: 1-based index of task_name column (8-колоночный блок или хвост 10-колоночного)."""
    off = start_col - 1
    vals = row_vals[off : off + 8]
    while len(vals) < 8:
        vals.append(None)

    return {
        "contract_number": contract_number,
        "etap_number": etap_number,
        "task_name": "" if vals[0] is None else str(vals[0]).strip(),
        "task_adenda": "" if vals[1] is None else str(vals[1]).strip(),
        "worker_name": "" if vals[2] is None else str(vals[2]).strip(),
        "task_start_date": vals[3],
        "task_end_date": vals[4],
        "task_comment": "" if vals[5] is None else str(vals[5]).strip(),
        "working_file": "" if vals[6] is None else str(vals[6]).strip(),
        "task_status": "" if vals[7] is None else str(vals[7]).strip(),
    }


def _parse_row_by_map(
    row_vals: List[Any],
    col_map: Dict[str, int],
    *,
    contract_number: str,
    etap_number: str,
) -> Dict[str, Any]:
    def cell(field: str) -> Any:
        c = col_map.get(field)
        if c is None or c < 1 or c > len(row_vals):
            return None
        return row_vals[c - 1]

    return {
        "contract_number": contract_number,
        "etap_number": etap_number,
        "task_name": "" if cell("task_name") is None else str(cell("task_name")).strip(),
        "task_adenda": "" if cell("task_adenda") is None else str(cell("task_adenda")).strip(),
        "worker_name": "" if cell("worker_name") is None else str(cell("worker_name")).strip(),
        "task_start_date": cell("task_start_date"),
        "task_end_date": cell("task_end_date"),
        "task_comment": "" if cell("task_comment") is None else str(cell("task_comment")).strip(),
        "working_file": "" if cell("working_file") is None else str(cell("working_file")).strip(),
        "task_status": "" if cell("task_status") is None else str(cell("task_status")).strip(),
    }


def parse_tasks_from_xlsx(
    *,
    file_bytes: bytes,
    expected_contract_number: str,
    expected_etap_number: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Парсинг проектного задания из Excel.
    Номер договора и этапа в файле не используются — подставляются expected_contract_number / expected_etap_number.
    Поддерживаются:
    - строка заголовков с именами столбцов (как в шаблоне, в т.ч. с колонками «Номер договора»/«Номер этапа» — они пропускаются);
    - 8 колонок подряд (только поля задачи);
    - устаревший формат из 10 колонок: первые два столбца игнорируются, данные в 3–10.
    """
    import io

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if not wb.sheetnames:
        return []
    ws = wb[wb.sheetnames[0]]

    etap_str = "" if expected_etap_number is None else str(expected_etap_number).strip()

    header_row_idx: Optional[int] = None
    col_map: Dict[str, int] = {}
    max_scan = min(ws.max_row, 50)
    max_col_scan = min(ws.max_column, 20)

    for r in range(1, max_scan + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col_scan + 1)]
        cmap = _build_column_map_from_header_row(row_vals)
        if "task_name" in cmap:
            header_row_idx = r
            col_map = cmap
            break

    tasks: List[Dict[str, Any]] = []

    if header_row_idx is not None:
        for r in range(header_row_idx + 1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col_scan + 1)]
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue
            t = _parse_row_by_map(
                row_vals,
                col_map,
                contract_number=expected_contract_number,
                etap_number=etap_str,
            )
            if not any(
                [
                    t["task_name"],
                    t["task_adenda"],
                    t["worker_name"],
                    t["task_comment"],
                    t["working_file"],
                    t["task_status"],
                    t["task_start_date"] not in (None, ""),
                    t["task_end_date"] not in (None, ""),
                ]
            ):
                continue
            tasks.append(t)
        return tasks

    # Fallback: без распознанной строки заголовков — ищем первую строку с данными
    data_start: Optional[int] = None
    for r in range(1, max_scan + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 11)]
        if any(v is not None and str(v).strip() != "" for v in row_vals):
            data_start = r
            break
    if data_start is None:
        return []

    first_row = [ws.cell(row=data_start, column=c).value for c in range(1, 11)]
    c0 = "" if first_row[0] is None else str(first_row[0]).strip()
    c1 = "" if first_row[1] is None else str(first_row[1]).strip()
    c2 = first_row[2] if len(first_row) > 2 else None
    c2_nonempty = c2 is not None and str(c2).strip() != ""
    empty12 = c0 == "" and c1 == ""

    # Шаблон с пустыми колонками 1–2 и данными с 3-й; либо заполненный старый формат 10 колонок
    if len(first_row) >= 10 and empty12 and c2_nonempty:
        start_col = 3
    elif len(first_row) >= 10 and c0 == expected_contract_number and (not etap_str or c1 == etap_str):
        start_col = 3
    else:
        start_col = 1

    for r in range(data_start, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 11)]
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        t = _parse_row_by_positions(
            row_vals,
            start_col=start_col,
            contract_number=expected_contract_number,
            etap_number=etap_str,
        )

        if not any(
            [
                t["task_name"],
                t["task_adenda"],
                t["worker_name"],
                t["task_comment"],
                t["working_file"],
                t["task_status"],
                t["task_start_date"] not in (None, ""),
                t["task_end_date"] not in (None, ""),
            ]
        ):
            continue
        tasks.append(t)

    return tasks
